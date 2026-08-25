"""
Independent re-implementation of the Ramal calculation engine, written
directly from the raw Excel formula text (Prediction!B30:Q45, C59:G76,
and 'Stihir Kundali'!B1:Q16) -- NOT derived from or copied out of the
TypeScript engine in ramal-app/src/lib/engines/. The point is to have a
second, independently-authored implementation to cross-check the shipped
app against, so a transcription mistake in one is unlikely to be silently
mirrored in the other.

Outputs one JSON object per line to stdout for every (draw, house, type)
combination, to be piped into scripts/oracle-diff.mts which recomputes each
case with the real shipped TypeScript engine and reports mismatches:

    python3 scripts/oracle.py exhaustive | npx tsx scripts/oracle-diff.mts

"exhaustive" covers the full 16^4 x 12 x 2 = 1,572,864 draw/house/type
space in a few seconds. Last run (2026-08-24): 0 mismatches.
"""
import json
import os
import sys
import openpyxl

SRC = os.path.join(os.path.dirname(__file__), "..", "..", "Ramal Calculation.xlsx")

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---- Load Stihir Kundali (16 canonical figures) ----
ws = wb["Stihir Kundali"]
FIGURES = {}  # id -> {"pattern": (4 symbols), "nature":..., "type":..., "timingNumber":...}
for col in range(2, 18):
    house = col - 1
    pattern = tuple(str(ws.cell(row=r, column=col).value) for r in range(2, 6))
    FIGURES[house] = {
        "pattern": pattern,
        "type": ws.cell(row=9, column=col).value.upper(),
        "nature": ws.cell(row=11, column=col).value.upper(),
        "timingNumber": int(ws.cell(row=16, column=col).value),
    }
PATTERN_TO_ID = {v["pattern"]: k for k, v in FIGURES.items()}

# ---- Load Timings sheet (16 blocks x 16 place entries) ----
ws = wb["Timings"]
TIMING_BLOCKS = {}  # timingNumber -> {place: (days,months,years)}
r = 2
while r <= ws.max_row:
    original_house = ws.cell(row=r, column=1).value
    if original_house is None:
        r += 1
        continue
    timing_number = int(ws.cell(row=r, column=2).value)
    places = [ws.cell(row=r, column=c).value for c in range(5, 21)]
    days = [ws.cell(row=r + 1, column=c).value or 0 for c in range(5, 21)]
    months = [ws.cell(row=r + 2, column=c).value or 0 for c in range(5, 21)]
    years = [ws.cell(row=r + 3, column=c).value or 0 for c in range(5, 21)]
    block = {}
    for i in range(16):
        block[int(places[i])] = (int(days[i]), int(months[i]), int(years[i]))
    TIMING_BLOCKS[timing_number] = block
    r += 4


def add_bit(a, b):
    return "-" if a == b else "0"


def add_figure(a, b):
    return tuple(add_bit(a[i], b[i]) for i in range(4))


def build_chart(mother_ids):
    """Places 1-16, per Prediction!B30:Q33 formulas (re-derived independently)."""
    m = [FIGURES[i]["pattern"] for i in mother_ids]
    chart = {1: m[0], 2: m[1], 3: m[2], 4: m[3]}
    # Places 5-8: row-wise transpose across mothers (Prediction!J30:M33)
    chart[5] = tuple(m[k][0] for k in range(4))
    chart[6] = tuple(m[k][1] for k in range(4))
    chart[7] = tuple(m[k][2] for k in range(4))
    chart[8] = tuple(m[k][3] for k in range(4))
    chart[9] = add_figure(chart[1], chart[2])
    chart[10] = add_figure(chart[3], chart[4])
    chart[11] = add_figure(chart[5], chart[6])
    chart[12] = add_figure(chart[7], chart[8])
    chart[13] = add_figure(chart[9], chart[10])
    chart[14] = add_figure(chart[11], chart[12])
    chart[15] = add_figure(chart[13], chart[14])
    chart[16] = add_figure(chart[15], chart[1])
    return chart


def guard_status(chart):
    """Prediction!C35 / C36 -- Place 15 only for 3-identical, Place 15 OR Place 1 for 4-identical."""
    def count(p, s):
        return sum(1 for x in p if x == s)

    p15, p1 = chart[15], chart[1]
    if count(p15, "0") == 3 or count(p15, "-") == 3:
        return "CALCULATION_ERROR"
    if count(p15, "0") == 4 or count(p15, "-") == 4 or count(p1, "0") == 4 or count(p1, "-") == 4:
        return "CANT_PREDICT_TODAY"
    return "OK"


def calculate_question(house, qtype, chart):
    """Prediction!B42:G45 -- exact literal formula, including the House-5 IF branch."""
    q_fig = chart[house]
    h1_fig = chart[1]  # C42:C45 is hardcoded to column Q (Place 1) regardless of house
    if house == 5:
        result = q_fig
    else:
        result = add_figure(q_fig, h1_fig)

    sthir_pattern = FIGURES[house]["pattern"]
    sthan_bali = result == sthir_pattern

    if sthan_bali:
        status = "YES"
    elif qtype == "AGAM" and result[0] == "-":
        status = "YES"
    elif qtype == "NIRGAM" and result[0] == "0":
        status = "YES"
    else:
        status = "NO"

    return {"questionHouseFigure": q_fig, "house1Figure": h1_fig, "resultFigure": result, "sthanBali": sthan_bali, "status": status}


def compute_timing(result_pattern, chart):
    """Prediction!C59:G76 -- Sthir match -> Timings Number -> scan CURRENT chart for
    every place whose pattern equals the result -> aggregate -> normalize."""
    fig_id = PATTERN_TO_ID.get(result_pattern)
    if fig_id is None:
        return {"timingNumber": -1, "matches": [], "totalDays": 0, "totalMonths": 0, "totalYears": 0, "unavailable": True}

    timing_number = FIGURES[fig_id]["timingNumber"]
    block = TIMING_BLOCKS.get(timing_number)
    if block is None:
        return {"timingNumber": timing_number, "matches": [], "totalDays": 0, "totalMonths": 0, "totalYears": 0, "unavailable": True}

    matches = []
    for place in range(1, 17):
        if chart[place] == result_pattern:
            d, mo, y = block.get(place, (0, 0, 0))
            matches.append({"place": place, "days": d, "months": mo, "years": y})

    sum_days = sum(m["days"] for m in matches)
    sum_months = sum(m["months"] for m in matches)
    sum_years = sum(m["years"] for m in matches)

    # E60/F60/G60 formulas exactly
    total_years = sum_years + ((sum_months + (sum_days // 30)) // 12)
    total_months = (sum_months + (sum_days // 30)) % 12
    total_days = sum_days % 30

    return {"timingNumber": timing_number, "matches": matches, "totalDays": total_days, "totalMonths": total_months, "totalYears": total_years, "unavailable": False}


ABJAD_VALUE = [1, 2, 3, 4]  # tez, vayu, jal, prithvi


def compute_quick_duration(result_pattern, short_timing):
    """Prediction!B90:F91 -- re-derived independently from the raw formulas
    (D90/E90/F90 for Normal, D91/E91/F91 for Short Duration). F90's unit
    lookup was originally broken (only resolved Day(s)/Week(s), houses
    9-16 fell through to "") but the owner fixed it in the workbook itself
    (2026-08-26) to check D95:D98 directly -- re-verified against the
    corrected live formula text before updating this to match. E91 (Short
    count) still excludes the tez/first-symbol contribution entirely --
    that bug was not corrected, still reproduced faithfully."""
    fig_id = PATTERN_TO_ID.get(result_pattern)
    if fig_id is None:
        return {"mode": "SHORT" if short_timing else "NORMAL", "sthirHouseId": None, "count": None, "unitLabel": ""}

    if not short_timing:
        count = sum(w for w, s in zip(ABJAD_VALUE, result_pattern) if s == "0")
        if 1 <= fig_id <= 4:
            unit = "Day(s)"
        elif 5 <= fig_id <= 8:
            unit = "Week(s)"
        elif 9 <= fig_id <= 12:
            unit = "Month(s)"
        else:
            unit = "Year(s)"  # 13-16, per Prediction!D95:D98
        return {"mode": "NORMAL", "sthirHouseId": fig_id, "count": count, "unitLabel": unit}

    # Bug 2: SUMIF(D86:D89,...) skips tez (D85) entirely.
    count = sum(w for w, s in zip([2, 3, 4], result_pattern[1:]) if s == "0")
    unit = "Minutes" if 1 <= fig_id <= 7 else "Hours"
    return {"mode": "SHORT", "sthirHouseId": fig_id, "count": count, "unitLabel": unit}


def full_case(mother_ids, house, qtype):
    chart = build_chart(mother_ids)
    status = guard_status(chart)
    judgement = calculate_question(house, qtype, chart)
    timing = compute_timing(judgement["resultFigure"], chart)
    qd_normal = compute_quick_duration(judgement["resultFigure"], False)
    qd_short = compute_quick_duration(judgement["resultFigure"], True)
    return {
        "draw": list(mother_ids),
        "house": house,
        "type": qtype,
        "guardStatus": status,
        "chart": {str(k): "".join(v) for k, v in chart.items()},
        "sthanBali": judgement["sthanBali"],
        "resultFigure": "".join(judgement["resultFigure"]),
        "status": judgement["status"],
        "timingNumber": timing["timingNumber"],
        "timingMatches": [m["place"] for m in timing["matches"]],
        "totalDays": timing["totalDays"],
        "totalMonths": timing["totalMonths"],
        "totalYears": timing["totalYears"],
        "quickDurationNormal": qd_normal,
        "quickDurationShort": qd_short,
    }


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "exhaustive"
    out = sys.stdout
    if mode == "exhaustive":
        # Full 16^4 draws x 12 houses x 2 types = 1,572,864 cases.
        for a in range(1, 17):
            for b in range(1, 17):
                for c in range(1, 17):
                    for d in range(1, 17):
                        chart = build_chart((a, b, c, d))
                        status = guard_status(chart)
                        for house in range(1, 13):
                            for qtype in ("AGAM", "NIRGAM"):
                                judgement = calculate_question(house, qtype, chart)
                                timing = compute_timing(judgement["resultFigure"], chart)
                                qd_normal = compute_quick_duration(judgement["resultFigure"], False)
                                qd_short = compute_quick_duration(judgement["resultFigure"], True)
                                rec = {
                                    "draw": [a, b, c, d],
                                    "house": house,
                                    "type": qtype,
                                    "guardStatus": status,
                                    "sthanBali": judgement["sthanBali"],
                                    "resultFigure": "".join(judgement["resultFigure"]),
                                    "status": judgement["status"],
                                    "timingNumber": timing["timingNumber"],
                                    "timingMatches": timing["matches"] and [m["place"] for m in timing["matches"]],
                                    "totalDays": timing["totalDays"],
                                    "totalMonths": timing["totalMonths"],
                                    "totalYears": timing["totalYears"],
                                    "quickDurationNormal": qd_normal,
                                    "quickDurationShort": qd_short,
                                }
                                out.write(json.dumps(rec, sort_keys=True) + "\n")
    elif mode == "benchmark":
        print(json.dumps(full_case((2, 8, 4, 9), 7, "NIRGAM"), indent=2, sort_keys=True))
